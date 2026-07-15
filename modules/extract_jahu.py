
# coding: UTF-8

# extract_jahu.py

import pdfplumber
import re
import os
import pandas as pd
from modules.read_ctms import read_ctms
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook

SCRIPT_DIR = os.path.dirname (os.path.abspath (__file__))
PROJECT_ROOT = os.path.abspath (os.path.join (SCRIPT_DIR, ".."))

def extract_jahu ():

    ctm_lines = read_ctms ()
    total_ctms = len (ctm_lines)
    index = 0

    while index < total_ctms:

        ctm_line = ctm_lines[index]

        try:

            pdf_path = os.path.join (PROJECT_ROOT, "JAHU", "PDFs", f"{ctm_line}.pdf")

            if not os.path.exists (pdf_path):

                print(f"AVISO: PDF NÃO LOCALIZADO PARA CÓDIGO CTM: {ctm_line} | IGNORANDO...")

            else:

                month_due = r'(\d{2}/\d{2}/\d{4})'
                date_count = 0
                venc_columns = ['MÊS_VENCIMENTO / ANO', 'VALOR', 'CORREÇÃO', 'MULTA', 'JUROS', 'DESCONTO', 'HONORÁRIOS', 'A PAGAR']
                venc_output_data = []

                with pdfplumber.open (pdf_path) as pdf_file:

                    for page in pdf_file.pages:

                        text = page.extract_text ()

                        if text:

                            lines = text.split ('\n')

                            lines = lines[7:-1]

                            for line in lines:

                                find_month_due = re.search (month_due, line)

                                if find_month_due:

                                    start_index = find_month_due.start ()
                                    n_index = line.find ('N', start_index)

                                    if n_index != -1:

                                        line_fragment = line[start_index:n_index]

                                    else:

                                        line_fragment = line[start_index:]

                                    row_data = line_fragment.strip().split ()

                                    row_data[0] = find_month_due.group(1)

                                    if len(row_data) < len(venc_columns):

                                        row_data += [''] * (len(venc_columns) - len(row_data))

                                    elif len(row_data) > len(venc_columns):

                                        row_data = row_data[:len(venc_columns)]

                                    venc_output_data.append (row_data)

                with pdfplumber.open (pdf_path) as pdf_file:

                    full_text = ''

                    for page in pdf_file.pages:

                        page_text = page.extract_text ()

                        if page_text:

                            full_text += page_text

                pattern_total_debts = r'.*Total da Dívida.*'
                matches_total_debts = re.findall (pattern_total_debts, full_text)

                total_columns = ['MÊS_VENCIMENTO / ANO', 'VALOR', 'CORREÇÃO', 'MULTA', 'JUROS', 'DESCONTO', 'HONORÁRIOS', 'A PAGAR']
                total_data = []

                total_years = get_years_by_total (full_text)
                year_index = 0

                if matches_total_debts:

                    for match_total_debt in matches_total_debts:

                        values = re.findall(r'\d{1,3}(?:\.\d{3})*,\d{2}', match_total_debt)

                        year = total_years[year_index] if year_index < len(total_years) else ''
                        year_index += 1

                        row = [year] + values

                        if len (row) < len (total_columns):
                            
                            row += [''] * (len (total_columns) - len (row))
                            
                        elif len (row) > len (total_columns):
                            
                            row = row[:len(total_columns)]

                        total_data.append (row)

                df_venc = pd.DataFrame (venc_output_data, columns = venc_columns) if venc_output_data else pd.DataFrame (columns = venc_columns)
                df_total = pd.DataFrame (total_data, columns = total_columns) if total_data else pd.DataFrame (columns = total_columns)

                df_final = pd.concat ([df_venc, df_total], ignore_index = True, sort = False)

                excel_out_dir = os.path.join (PROJECT_ROOT, "JAHU")
                excel_out_file = os.path.join (excel_out_dir, f"{ctm_line}.xlsx")

                df_final.to_excel (excel_out_file, index = False)

                print (f"CÓDIGO CTM: {ctm_line} PROCESSADO COM SUCESSO! ({index + 1}/{total_ctms})")

        except Exception as e:

            print (f"ERRO AO PROCESSAR CÓDIGO CTM: {ctm_line} | {str(e)}")
            print (f"CONTINUANDO PARA O PRÓXIMO CÓDIGO CTM...")

        finally:

            index += 1

    print (f"\nPROCESSAMENTO EXTRACT_JAHU CONCLUÍDO! TOTAL: {total_ctms} CTM'S")

def get_years_by_total (text):

    year_positions = [(match.start(), match.group(1)) for match in re.finditer(r'Exerc[ií]cio\s*:\s*(\d{4})', text, flags=re.I)]
    total_positions = [match.start() for match in re.finditer(r'Total da D[ií]vida', text, flags=re.I)]
    years, year_index = [], 0

    for total_pos in total_positions:

        while year_index + 1 < len(year_positions) and year_positions[year_index + 1][0] <= total_pos:

            year_index += 1

        years.append (year_positions[year_index][1] if year_positions and year_positions[year_index][0] <= total_pos else "")

    return years

def _find_valor_header_position (worksheet):
    
    for row_index in range (1, worksheet.max_row + 1):
        
        for col_index in range (1, worksheet.max_column + 1):
            
            cell_value = worksheet.cell (row_index, col_index).value
            
            if isinstance (cell_value, str) and cell_value.strip ().upper () == "VALOR":
                
                return row_index, col_index
                
    return None, None

_pos_header_valor = _find_valor_header_position

def add_year_by_exercise (base_dir=None, excels_subdir=("JAHU",), pdfs_subdir=("JAHU","PDFs"), sheet_name=None):
    
    base_dir = base_dir or PROJECT_ROOT
    excel_dir = os.path.join (base_dir, *excels_subdir)
    pdf_dir = os.path.join (base_dir, *pdfs_subdir)

    files = [f for f in os.listdir(excel_dir) if f.lower().endswith(".xlsx") and f not in ["CTM'S_STATUS.xlsx", "FINAL_CTM.xlsx"]]
    total_files = len (files)
    index = 0
    
    while index < total_files:
        
        file_name = files[index]
        
        try:
            
            xlsx_path = os.path.join (excel_dir, file_name)
            workbook = load_workbook (xlsx_path)
            worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            header_row, col_valor = _find_valor_header_position (worksheet)
            
            if not header_row or not col_valor:
                
                workbook.save(xlsx_path)
                
                print(f"AVISO: HEADER 'VALOR' NÃO ENCONTRADO EM {file_name} | PULANDO...")
                
            else:

                worksheet.insert_cols(col_valor)
                worksheet.cell(header_row, col_valor).value = "ANO"
                worksheet.cell(header_row, col_valor).font = Font(bold=True)
                worksheet.cell(header_row, col_valor).alignment = Alignment(horizontal="center", vertical="center")

                ctm = os.path.splitext (file_name)[0]
                pdf_path = os.path.join (pdf_dir, f"{ctm}.pdf")
                text_content = ""
                
                if os.path.exists (pdf_path):
                    
                    with pdfplumber.open (pdf_path) as pdf:
                        
                        for page in pdf.pages:
                            
                            page_text = page.extract_text ()
                            
                            if page_text:
                                
                                text_content += page_text

                years = get_years_by_total (text_content)
                
                for row_offset, year in enumerate (years, start=1):
                    
                    worksheet.cell (header_row + row_offset, col_valor).value = year

                workbook.save (xlsx_path)
                
                print (f"YEARS ADICIONADOS EM {file_name} ({index + 1}/{total_files})")
        
        except Exception as e:
            
            print(f"ERRO AO PROCESSAR {file_name}: {str(e)}")
            print(f"CONTINUANDO PARA O PRÓXIMO ARQUIVO...")
        
        finally:
            
            index += 1
    
    print (f"\nPROCESSAMENTO ADD_YEAR_BY_EXERCISE CONCLUÍDO! TOTAL: {total_files} ARQUIVOS")

add_ano_por_exercicio = add_year_by_exercise

def _parse_address_block (text):
    
    matches = re.findall (r'Endereço[\s\S]*?(?=\s*Distrito)', text, flags=re.IGNORECASE)
    cleaned_block = (matches[0] if matches else "").replace("X", "").strip()
    cleaned_block = re.sub(r'(ENDEREÇO|BAIRRO|CEP)', '', cleaned_block, flags=re.IGNORECASE).strip()
    zip_match = re.search (r'(\d{8})', cleaned_block)
    zipcode = zip_match.group (1) if zip_match else "CEP NÃO ENCONTRADO!"
    cleaned_block = cleaned_block.replace (zipcode, '').strip() if zip_match else cleaned_block
    
    if ',' in cleaned_block:
        
        street_part, after_comma = cleaned_block.split (',', 1)
        number_match = re.search (r'\d+', after_comma)
        
        if number_match:
            
            address = street_part.strip () + ', ' + number_match.group (0)
            neighborhood = after_comma.replace (number_match.group (0), '').strip ()
            neighborhood = ' '.join (word for word in neighborhood.split () if not word.isdigit ())
            
        else:
            
            address, neighborhood = street_part.strip (), after_comma.strip ()
    else:
        
        address, neighborhood = cleaned_block, "BAIRRO NÃO ENCONTRADO!"
        
    return address or "ENDEREÇO NÃO ENCONTRADO!", neighborhood, zipcode

def _parse_taxpayer (text: str) -> str:
    
    match = re.search (r"Contribuinte:(.*?)(?:CPF/CNPJ:|Origem:)", text, flags=re.DOTALL | re.IGNORECASE)
    
    if match:
        
        return match.group (1).strip ()
    
    return ""

def add_title_and_address (base_dir=None, excels_subdir=("JAHU",), pdfs_subdir=("JAHU","PDFs"), sheet_name=None, title_prefix="CÓDIGO CTM: "):

    base_dir = base_dir or PROJECT_ROOT
    excel_dir = os.path.join (base_dir, *excels_subdir)
    pdf_dir = os.path.join (base_dir, *pdfs_subdir)

    files = [f for f in os.listdir(excel_dir) if f.lower().endswith(".xlsx") and f not in ["CTM'S_STATUS.xlsx", "FINAL_CTM.xlsx"]]
    total_files = len(files)
    index = 0
    
    while index < total_files:
        
        file_name = files[index]
        
        try:
            
            xlsx_path = os.path.join (excel_dir, file_name)
            workbook = load_workbook (xlsx_path)
            worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            last_col_letter = get_column_letter (worksheet.max_column)

            worksheet.insert_rows (1)
            
            for merge_range in list (worksheet.merged_cells.ranges):
                
                if merge_range.min_row == 1 and merge_range.max_row == 1:
                    
                    worksheet.unmerge_cells (str (merge_range))
                        
            worksheet.merge_cells (f"A1:{last_col_letter}1")
            
            ctm = os.path.splitext (file_name)[0]
            
            worksheet["A1"].value = f"{title_prefix}{ctm}"
            worksheet["A1"].alignment = Alignment (horizontal="center", vertical="center")
            worksheet["A1"].font = Font (bold=True, size=12)
            
            worksheet.row_dimensions[1].height = 22

            has_addr_block = isinstance (worksheet["A2"].value, str) and worksheet["A2"].value.upper ().startswith ("ENDEREÇO:")
            
            if not has_addr_block:
                
                worksheet.insert_rows(2, amount=3)

            pdf_path = os.path.join(pdf_dir, f"{ctm}.pdf")
            text_content = ""

            if os.path.exists (pdf_path):
                
                with pdfplumber.open (pdf_path) as pdf:
                    
                    for page in pdf.pages:
                        
                        page_text = page.extract_text ()
                        
                        if page_text:
                            
                            text_content += page_text

            taxpayer = _parse_taxpayer (text_content)
            address, neighborhood, zipcode = _parse_address_block (text_content)

            rows_data = [
                (2, "ENDEREÇO", address),
                (3, "BAIRRO", neighborhood),
                (4, "CEP", zipcode),
            ]

            for row_index, label, value in rows_data:

                for merge_range in list (worksheet.merged_cells.ranges):
                    
                    if merge_range.min_row == row_index and merge_range.max_row == row_index:
                        
                        worksheet.unmerge_cells (str (merge_range))

                worksheet.merge_cells (f"A{row_index}:{last_col_letter}{row_index}")
                
                cell = worksheet[f"A{row_index}"]
                cell.value = f"{label}: {value}"
                cell.alignment = Alignment (horizontal="left", vertical="center")
                cell.font = Font (bold=True)

            if taxpayer:
                
                contrib_row = 5

                worksheet.insert_rows (contrib_row, amount=1)

                for merge_range in list (worksheet.merged_cells.ranges):
                    
                    if merge_range.min_row == contrib_row and merge_range.max_row == contrib_row:
                        
                        worksheet.unmerge_cells (str (merge_range))

                worksheet.merge_cells (f"A{contrib_row}:{last_col_letter}{contrib_row}")
                
                cell = worksheet[f"A{contrib_row}"]
                cell.value = f"CONTRIBUINTE: {taxpayer}"
                cell.alignment = Alignment (horizontal="left", vertical="center")
                cell.font = Font (bold=True)

                header_row = contrib_row + 1

            else:
                
                header_row = 5

            for cell in worksheet[header_row]:
                
                cell.font = Font (bold=True)
                cell.alignment = Alignment (horizontal="center", vertical="center")

            workbook.save (xlsx_path)
            
            print (f"TÍTULO E ENDEREÇO ADICIONADOS EM {file_name} ({index + 1}/{total_files})")
        
        except Exception as e:
            
            print (f"ERRO AO PROCESSAR {file_name}: {str(e)}")
            print ("CONTINUANDO PARA O PRÓXIMO ARQUIVO...")
        
        finally:
            
            index += 1
    
    print (f"\nPROCESSAMENTO ADD_TITLE_AND_ADDRESS CONCLUÍDO! TOTAL: {total_files} ARQUIVOS")

def exclude_column_year ():
    
    excel_dir = os.path.join (PROJECT_ROOT, "JAHU")
    ctms = read_ctms ()

    for ctm in ctms:

        excel_path = os.path.join (excel_dir, f"{ctm}.xlsx")

        if os.path.exists (excel_path):

            wb = load_workbook (excel_path)
            ws = wb.active
            ws.delete_cols (2)
            wb.save (excel_path)

def merge_excels_into_final (excel_dir=None, final_file_name="FINAL_CTM.xlsx"):

    excel_dir = excel_dir or os.path.join (PROJECT_ROOT, "JAHU")
    final_file_path = os.path.join (excel_dir, final_file_name)

    if os.path.exists (final_file_path):
        
        os.remove (final_file_path)

    final_wb = Workbook ()
    final_ws = final_wb.active
    final_ws.title = "CTM'S_CODES"

    row_offset = 1

    date_pattern = re.compile(r'\b\d{2}/\d{2}/\d{4}\b')

    for file_name in os.listdir (excel_dir):
        
        if file_name == "CTM'S_STATUS.xlsx" or file_name == final_file_name:
            
            continue
        
        if file_name.lower ().endswith (".xlsx"):
            
            file_path = os.path.join (excel_dir, file_name)
            
            workbook = load_workbook (file_path, data_only=True)
            sheet = workbook.active

            for row in sheet.iter_rows (values_only=True):
                
                if any (value and isinstance (value, str) and date_pattern.search (value) for value in row):
                    
                    continue
                
                for col_index, value in enumerate (row):
                    
                    final_ws.cell (row=row_offset, column=col_index + 1, value=value)
                    
                row_offset += 1

            row_offset += 1

    final_wb.save (final_file_path)
    
    print (f'ARQUIVOS MESCLADOS COM SUCESSO!')
