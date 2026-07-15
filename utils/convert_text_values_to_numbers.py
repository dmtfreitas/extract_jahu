
# coding: UTF-8

# convert_text_values_to_numbers.py

import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def convert_text_values_to_numbers ():
    
    get_cwd = os.getcwd ()
    
    folder_path = os.path.dirname (os.path.abspath (__file__))
    return_folder = os.path.abspath (os.path.join (folder_path, '..'))
    jahu_folder = os.path.join (get_cwd, 'JAHU')

    target_headers = ['VALOR', 'CORREÇÃO', 'MULTA', 'JUROS', 'DESCONTO', 'HONORÁRIOS', 'A PAGAR']

    excel_files = [f for f in os.listdir (jahu_folder) if f.lower ().endswith (".xlsx")]

    for file_name in excel_files:
        
        file_path = os.path.join (jahu_folder, file_name)

        try:
            
            workbook = load_workbook (file_path)
            sheet = workbook.active

            header_row = None
            detected_columns = {}

            for row in sheet.iter_rows (min_row=1, max_row=30):
                
                for cell in row:
                    
                    if isinstance (cell.value, str):
                        
                        header_name = cell.value.strip ().upper ()
                        
                        if header_name in target_headers:
                            
                            header_row = cell.row
                            detected_columns[header_name] = cell.column
                            
                if header_row:
                    
                    break

            if not header_row:
                
                print (f'[WARNING] HEADER NOT FOUND IN: {file_name}')
                continue

            for header_name, col_idx in detected_columns.items ():
                
                col_letter = get_column_letter (col_idx)

                for row_num in range (header_row + 1, sheet.max_row + 1):
                    
                    cell = sheet[f"{col_letter}{row_num}"]
                    raw_value = cell.value

                    if isinstance (raw_value, str):
                        
                        cleaned = raw_value.strip ().upper ()
                        cleaned = cleaned.replace ('R$', '').replace(' ', '')
                        cleaned = cleaned.replace ('.', '').replace(',', '.')

                        try:
                            
                            number = float (cleaned)
                            cell.value = number
                            cell.number_format = '#,##0.00'
                            
                        except:
                            
                            pass

            workbook.save (file_path)

        except Exception as error:
            
            print(f'ERROR IN FILE {file_name}: {error}')

